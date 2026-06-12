"""Multi-sheet variant of seo_quality_pipeline (task #173, 2026-05-25).

The MultiSheet template has 5 sheets (Batch_1..Batch_5) of 10k rows
each, 7 columns (no Stock Status, no 'Estimated' prefix). This script
processes ALL sheets, runs the same title-cleanup + dedupe + SEO build
pipeline, and writes per-batch output xlsx + a combined dedupe map.

Output naming:
  Daraz_Style_MultiSheet_Template.cleaned.batch_N.xlsx  (per batch)
  Daraz_Style_MultiSheet_Template.dedupe_map.csv        (combined)
  Daraz_Style_MultiSheet_Template.sample_50.csv         (first 50 rows)

Reuses the audience matrix + title-rewrite logic from
seo_quality_pipeline (imported); only the IO layer differs.
"""
from __future__ import annotations

import argparse
import csv
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl

from app.modules.seo.autogen import build_product_seo
from scripts.seo_quality_pipeline import (
    _dedupe_key,
    _extract_audience_from_title,
    _NEW_COLS as _BASE_NEW_COLS,
    build_jsonld,
    clean_seed_title,
    slugify,
    stock_to_availability,
)

_SITE_BASE = "https://hypershop.com.bd"
# MultiSheet headers — 7 cols (no Stock Status)
_HEADER = (
    "SKU", "Category", "Brand", "Core Product", "Specification",
    "Price (BDT)", "SEO Title",
)
# Output keeps the unified 8-col base + 12 SEO cols
_OUT_HEADER = (
    "SKU", "Category", "Brand", "Core Product", "Specification",
    "Stock Status", "Estimated Price (BDT)", "SEO Optimized Title",
)
_NEW_COLS = _BASE_NEW_COLS  # same 12 SEO columns


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--in", dest="in_path", required=True)
    parser.add_argument("--out-prefix", dest="out_prefix", required=True,
                        help="Base path; per-batch xlsx appends .batch_N.xlsx")
    parser.add_argument("--dedupe-map", dest="dedupe_map_path", required=True)
    parser.add_argument("--sample-csv", dest="sample_csv_path", required=True)
    parser.add_argument("--sample-rows", type=int, default=50)
    parser.add_argument("--site-base", default=_SITE_BASE)
    parser.add_argument("--limit-per-batch", type=int, default=0,
                        help="Process only first N rows per batch (0 = all)")
    args = parser.parse_args()

    in_path = Path(args.in_path)
    if not in_path.exists():
        print(f"input not found: {in_path}", file=sys.stderr)
        return 2

    print(f"reading: {in_path}")
    wb_in = openpyxl.load_workbook(in_path, read_only=True, data_only=True)
    sheets = [s for s in wb_in.sheetnames if s.startswith("Batch_")]
    if not sheets:
        print(f"no Batch_* sheets found in {wb_in.sheetnames}", file=sys.stderr)
        return 2

    # Dedupe is GLOBAL across all batches — same (brand, core, spec, audience)
    # in Batch_3 and Batch_5 should collapse to one canonical URL.
    seen_slugs: dict[str, int] = {}
    global_groups: dict[tuple, str] = {}  # dedupe_key -> canonical slug
    dedupe_csv_rows: list[tuple[str, str]] = []
    sample_rows: list[list[Any]] = []
    n_total = 0
    n_fixed_total = 0
    n_canonical_total = 0
    n_redirect_total = 0

    for sheet_name in sheets:
        ws = wb_in[sheet_name]
        batch_num = sheet_name.split("_")[1]
        out_path = Path(f"{args.out_prefix}.batch_{batch_num}.xlsx")

        print(f"\n=== {sheet_name} -> {out_path.name} ===")
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter)
        if tuple(str(c or "") for c in header[: len(_HEADER)]) != _HEADER:
            print(f"  WARN: header mismatch {header}", file=sys.stderr)

        wb_out = openpyxl.Workbook(write_only=True)
        ws_out = wb_out.create_sheet(title=f"batch_{batch_num}_cleaned")
        ws_out.append(list(_OUT_HEADER) + list(_NEW_COLS))

        n_batch = 0
        n_fixed = 0
        n_canon = 0
        n_redir = 0
        for row in rows_iter:
            if args.limit_per_batch and n_batch >= args.limit_per_batch:
                break
            if not row or not row[0]:
                continue
            sku = str(row[0] or "").strip()
            category = str(row[1] or "").strip()
            brand = str(row[2] or "").strip()
            core = str(row[3] or "").strip()
            spec = str(row[4] or "").strip()
            price = row[5]
            seed_title = str(row[6] or "").strip()
            if not sku or not core:
                continue

            cleaned_title, was_fixed = clean_seed_title(seed_title, core)
            if was_fixed:
                n_fixed += 1

            audience = _extract_audience_from_title(cleaned_title)
            key = _dedupe_key(brand, core, spec, audience)

            display_name = cleaned_title
            base_slug = slugify(display_name) or slugify(sku.lower())
            n = seen_slugs.get(base_slug, 0)
            slug = base_slug if n == 0 else f"{base_slug}-{n + 1}"
            seen_slugs[base_slug] = n + 1
            canonical_url = f"{args.site_base.rstrip('/')}/product/{slug}"

            if key in global_groups:
                canonical_slug = global_groups[key]
                is_canonical = False
                n_redir += 1
            else:
                global_groups[key] = slug
                canonical_slug = slug
                is_canonical = True
                n_canon += 1

            seo = build_product_seo(
                name=display_name,
                brand=brand or None,
                category=category or None,
                short_description=spec or None,
            )
            try:
                price_minor = int(round(float(price or 0) * 100))
            except (TypeError, ValueError):
                price_minor = 0
            jsonld = build_jsonld(
                sku=sku, name=display_name,
                brand=brand or None, category=category or None,
                description=seo["en"]["description"],
                price_minor=price_minor,
                # MultiSheet has no Stock Status — default to InStock
                availability=stock_to_availability("In Stock"),
                url=canonical_url,
            )

            out_row = [
                sku, category, brand, core, spec,
                "In Stock", price, cleaned_title,
                slug, canonical_url, canonical_slug, is_canonical,
                seo["en"]["title"], seo["en"]["description"], seo["en"]["keywords"],
                seo["bn"]["title"], seo["bn"]["description"], seo["bn"]["keywords"],
                jsonld,
                was_fixed,
            ]
            ws_out.append(out_row)
            dedupe_csv_rows.append((slug, canonical_slug))
            if len(sample_rows) < args.sample_rows:
                sample_rows.append(out_row)
            n_batch += 1

            if n_batch % 2000 == 0:
                print(f"  {n_batch} rows...")

        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb_out.save(out_path)
        print(f"  batch wrote {n_batch} rows  fixed={n_fixed}  "
              f"canonical={n_canon}  redirects={n_redir}")
        n_total += n_batch
        n_fixed_total += n_fixed
        n_canonical_total += n_canon
        n_redirect_total += n_redir

    # Combined dedupe map
    dedupe_map_path = Path(args.dedupe_map_path)
    with dedupe_map_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["slug", "canonical_slug", "is_canonical"])
        for slug, canon in dedupe_csv_rows:
            w.writerow([slug, canon, slug == canon])
    print(f"\nwrote dedupe map: {dedupe_map_path}")

    # Sample CSV
    sample_path = Path(args.sample_csv_path)
    with sample_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(list(_OUT_HEADER) + list(_NEW_COLS))
        for row in sample_rows:
            w.writerow(row)
    print(f"wrote sample csv: {sample_path}  ({len(sample_rows)} rows)")

    print()
    print("=" * 60)
    print(f"  total rows           : {n_total}")
    print(f"  audience fixed       : {n_fixed_total} "
          f"({100.0 * n_fixed_total / max(1, n_total):.1f}%)")
    print(f"  canonical (global)   : {n_canonical_total}")
    print(f"  redirects (global)   : {n_redirect_total} "
          f"({100.0 * n_redirect_total / max(1, n_total):.1f}%)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
