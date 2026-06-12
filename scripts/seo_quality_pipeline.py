"""4-step SEO quality pipeline (task #172, 2026-05-25).

Inputs: the 10k Daraz xlsx + the prior with_seo.xlsx output.
Outputs:
  Daraz_Style_10k_SEO_Items.cleaned.xlsx   — title-cleaned + rebuilt SEO
  Daraz_Style_10k_SEO_Items.dedupe_map.csv — slug -> canonical_slug
  Daraz_Style_10k_SEO_Items.sample_50.csv  — preview of top 50 cleaned rows

Pipeline:
  STEP 1: detect product<->audience mismatches in seed title, rewrite
          audience token from a product-validity matrix. Eg "Core i5
          Laptop ... for Babies" -> "... for Office".
  STEP 2: regenerate slug + canonical + EN/BN meta + JSON-LD on the
          fixed display name using the same build_product_seo builder.
  STEP 3: dedupe pass. Group rows by (brand, product, spec, audience)
          tuple. First in group keeps the canonical slug; rest get
          ``canonical_slug`` pointer for a 301 redirect map.
  STEP 4: write 50-row sample CSV for preview.

The DB ingest step is separate (seo_bulk_ingest_to_db.py) since it
needs the backend running; this script is offline + idempotent.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl

from app.modules.seo.autogen import build_product_seo


# ---------------------------------------------------------------------------
# STEP 1 — product <-> audience validity matrix
# ---------------------------------------------------------------------------
# Each product key has an ordered list of valid audiences. The cleanup
# picks the first valid audience when the seed title's audience is not
# in the list, OR (when the seed audience IS valid) keeps it as-is.
# Order matters: the first entry is the natural-fit fallback.
_PRODUCT_AUDIENCE: dict[str, list[str]] = {
    # baby-only goods
    "Baby Clothes": ["Babies"],
    "Baby Food Formula": ["Babies"],
    "Baby Wipes": ["Babies"],
    "Baby Blanket": ["Babies", "Home"],
    "Feeding Bottle": ["Babies"],
    "Diapers": ["Babies"],
    "Newborn Care Gift Box": ["Babies"],
    # womens fashion
    "Silk Sari": ["Women", "Home"],
    "Designer Blouse": ["Women"],
    "Georgette Kurti": ["Women"],
    "Gold Plated Jewelry": ["Women", "Men"],
    # mens fashion
    "Traditional Panjabi": ["Men"],
    "Slim Fit Denim": ["Men", "Women"],
    "Cotton T-Shirt": ["Men", "Women", "Babies", "Home"],
    # universal-ish
    "Running Sneakers": ["Men", "Women", "Home"],
    # electronics / appliances — never for babies
    "Core i5 Laptop": ["Office", "Home", "Men", "Women"],
    "Gaming Keyboard": ["Office", "Home", "Men"],
    "Wireless Mouse": ["Office", "Home", "Men", "Women"],
    "Wireless Earbuds": ["Home", "Office", "Men", "Women"],
    "5G Smartphone 128GB": ["Men", "Women", "Home", "Office"],
    "Smart TV 43 Inch": ["Home"],
    "Inverter AC 1.5 Ton": ["Home", "Office"],
    "Frost Refrigerator 500L": ["Home"],
    "Washing Machine": ["Home"],
    "CCTV Camera": ["Home", "Office"],
    "4K Drone": ["Men", "Office", "Home"],
    "Power Bank 20000mAh": ["Office", "Men", "Women", "Home"],
    "Type-C Cable": ["Office", "Home", "Men", "Women"],
    "3D Acrylic Wall Clock": ["Home", "Office"],
}

# Audience tokens that may appear in seed titles. The cleanup only
# rewrites these tokens — anything else passes through unchanged.
_KNOWN_AUDIENCES = {"Babies", "Men", "Women", "Home", "Office"}

# Adjective bank — used to detect the seed title's prefix when we
# need to recompose a title from scratch.
_KNOWN_ADJECTIVES = {
    "Best", "Premium", "Top Rated", "Trending", "Exclusive",
    "Affordable", "Buy",
}


def _extract_audience_from_title(title: str) -> str | None:
    """Pull the audience token from a seed title like '... for Babies'."""
    m = re.search(r"\bfor\s+([A-Z][a-zA-Z]+)$", title.strip())
    if not m:
        return None
    aud = m.group(1)
    return aud if aud in _KNOWN_AUDIENCES else None


def _rewrite_audience(title: str, valid_audiences: list[str]) -> str:
    """Swap the audience token at the end of the title for the first
    valid audience when the current one isn't allowed. Preserves
    everything before 'for X'."""
    current = _extract_audience_from_title(title)
    if current is None:
        return title
    if current in valid_audiences:
        return title
    target = valid_audiences[0]
    return re.sub(
        r"\bfor\s+[A-Z][a-zA-Z]+$",
        f"for {target}",
        title.strip(),
    )


def _resolve_product_key(core_product: str) -> str | None:
    """Match the catalog Core Product field against our audience matrix.
    Partial-match the longest key that's a substring of the core name —
    handles seller-supplied prefixes like 'Newborn Care Gift Box'."""
    if not core_product:
        return None
    cp = core_product.strip()
    # Direct hit
    if cp in _PRODUCT_AUDIENCE:
        return cp
    # Longest-substring fallback
    matches = [k for k in _PRODUCT_AUDIENCE if k in cp or cp in k]
    if not matches:
        return None
    matches.sort(key=len, reverse=True)
    return matches[0]


def clean_seed_title(seed: str, core_product: str) -> tuple[str, bool]:
    """Returns (cleaned_title, was_changed)."""
    key = _resolve_product_key(core_product)
    if not key:
        return seed, False
    valid = _PRODUCT_AUDIENCE[key]
    fixed = _rewrite_audience(seed, valid)
    return fixed, fixed != seed


# ---------------------------------------------------------------------------
# STEP 2/3 — SEO rebuild + dedupe (mirrors seo_bulk_from_xlsx layout)
# ---------------------------------------------------------------------------
_SITE_BASE = "https://hypershop.com.bd"
_SHEET_IN = "10k_Database"
_HEADER = (
    "SKU", "Category", "Brand", "Core Product", "Specification",
    "Stock Status", "Estimated Price (BDT)", "SEO Optimized Title",
)
_NEW_COLS = (
    "slug", "canonical_url", "canonical_slug",  # canonical_slug = slug when row IS canonical, else target
    "is_canonical",  # bool
    "meta_title_en", "meta_desc_en", "meta_keywords_en",
    "meta_title_bn", "meta_desc_bn", "meta_keywords_bn",
    "jsonld_product",
    "audience_fixed",  # bool — STEP 1 changed it
)


def slugify(text: str) -> str:
    text = (text or "").lower().strip()
    text = re.sub(r"[^a-z0-9\s-]", "", text)
    text = re.sub(r"\s+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text[:80]


def stock_to_availability(stock: str) -> str:
    s = (stock or "").strip().lower()
    if s in ("low stock", "limited"):
        return "https://schema.org/LimitedAvailability"
    if s in ("out of stock", "sold out"):
        return "https://schema.org/OutOfStock"
    return "https://schema.org/InStock"


def build_jsonld(
    *, sku: str, name: str, brand: str | None, category: str | None,
    description: str, price_minor: int, availability: str, url: str,
) -> str:
    obj: dict[str, Any] = {
        "@context": "https://schema.org",
        "@type": "Product",
        "name": name,
        "sku": sku,
        "description": description,
        "url": url,
        "offers": {
            "@type": "Offer",
            "price": f"{price_minor / 100:.2f}",
            "priceCurrency": "BDT",
            "availability": availability,
            "url": url,
            "seller": {"@type": "Organization", "name": "Hypershop BD"},
        },
    }
    if brand:
        obj["brand"] = {"@type": "Brand", "name": brand}
    if category:
        obj["category"] = category
    return json.dumps(obj, ensure_ascii=False, separators=(",", ":"))


def _dedupe_key(brand: str, core: str, spec: str, audience: str | None) -> tuple:
    """Group by (brand, core_product, spec, audience). When two rows
    share this 4-tuple they describe the same SKU semantically — only
    one canonical URL should point at it."""
    return (
        brand.strip().lower(),
        core.strip().lower(),
        spec.strip().lower(),
        (audience or "").lower(),
    )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--in", dest="in_path", required=True)
    parser.add_argument("--out", dest="out_path", required=True)
    parser.add_argument("--dedupe-map", dest="dedupe_map_path", required=True)
    parser.add_argument("--sample-csv", dest="sample_csv_path", required=True)
    parser.add_argument("--sample-rows", type=int, default=50)
    parser.add_argument("--site-base", default=_SITE_BASE)
    parser.add_argument("--limit", type=int, default=0)
    args = parser.parse_args()

    in_path = Path(args.in_path)
    if not in_path.exists():
        print(f"input not found: {in_path}", file=sys.stderr)
        return 2

    print(f"reading: {in_path}")
    wb_in = openpyxl.load_workbook(in_path, read_only=True, data_only=True)
    ws_in = wb_in[_SHEET_IN]

    rows_iter = ws_in.iter_rows(values_only=True)
    header = next(rows_iter)
    if tuple(str(c or "") for c in header[: len(_HEADER)]) != _HEADER:
        print(f"WARN: header mismatch: {header}", file=sys.stderr)

    # ---------- first pass: clean titles + collect rows ----------
    cleaned_rows: list[dict[str, Any]] = []
    n_audience_fixed = 0
    for idx, row in enumerate(rows_iter, start=1):
        if args.limit and len(cleaned_rows) >= args.limit:
            break
        sku = str(row[0] or "").strip()
        category = str(row[1] or "").strip()
        brand = str(row[2] or "").strip()
        core = str(row[3] or "").strip()
        spec = str(row[4] or "").strip()
        stock = str(row[5] or "").strip()
        price = row[6]
        seed_title = str(row[7] or "").strip()

        if not sku or not core:
            continue

        cleaned_title, was_fixed = clean_seed_title(seed_title, core)
        if was_fixed:
            n_audience_fixed += 1

        cleaned_rows.append({
            "sku": sku, "category": category, "brand": brand,
            "core": core, "spec": spec, "stock": stock, "price": price,
            "seed_title": seed_title, "cleaned_title": cleaned_title,
            "was_fixed": was_fixed,
        })

    print(f"step 1: {len(cleaned_rows)} rows read, "
          f"{n_audience_fixed} audience tokens rewritten")

    # ---------- second pass: dedupe groups ----------
    groups: dict[tuple, list[int]] = defaultdict(list)
    for i, r in enumerate(cleaned_rows):
        audience = _extract_audience_from_title(r["cleaned_title"])
        key = _dedupe_key(r["brand"], r["core"], r["spec"], audience)
        groups[key].append(i)

    canonical_of: dict[int, int] = {}
    for key, members in groups.items():
        canonical_idx = members[0]
        for m in members:
            canonical_of[m] = canonical_idx

    n_canonical = len(groups)
    n_redirects = len(cleaned_rows) - n_canonical
    print(f"step 2: {n_canonical} canonical, {n_redirects} redirects in dedupe map")

    # ---------- third pass: build full SEO + write outputs ----------
    wb_out = openpyxl.Workbook(write_only=True)
    ws_out = wb_out.create_sheet(title="10k_cleaned")
    ws_out.append(list(_HEADER) + list(_NEW_COLS))

    seen_slugs: dict[str, int] = {}
    final_rows: list[list[Any]] = []  # held for sample CSV write
    dedupe_csv_rows: list[tuple[str, str]] = []  # (slug, canonical_slug)

    for i, r in enumerate(cleaned_rows):
        display_name = r["cleaned_title"]
        if len(display_name) < len(r["core"]):
            display_name = f"{r['brand']} {r['core']}".strip() or r["core"]

        base_slug = slugify(display_name) or slugify(r["sku"].lower())
        n = seen_slugs.get(base_slug, 0)
        slug = base_slug if n == 0 else f"{base_slug}-{n + 1}"
        seen_slugs[base_slug] = n + 1
        canonical_url = f"{args.site_base.rstrip('/')}/product/{slug}"

        canonical_idx = canonical_of[i]
        is_canonical = (canonical_idx == i)
        # For the canonical_slug column we need the canonical row's slug.
        # Resolve lazily — final_rows[canonical_idx][8] holds it once we
        # emit that row. Since dedupe-key first-member is always the
        # earliest index, canonical_idx <= i, so it's already in final_rows.
        if is_canonical:
            canonical_slug = slug
        else:
            canonical_slug = final_rows[canonical_idx][8]  # col index of "slug"

        seo = build_product_seo(
            name=display_name,
            brand=r["brand"] or None,
            category=r["category"] or None,
            short_description=r["spec"] or None,
        )

        try:
            price_minor = int(round(float(r["price"] or 0) * 100))
        except (TypeError, ValueError):
            price_minor = 0
        jsonld = build_jsonld(
            sku=r["sku"], name=display_name,
            brand=r["brand"] or None, category=r["category"] or None,
            description=seo["en"]["description"],
            price_minor=price_minor,
            availability=stock_to_availability(r["stock"]),
            url=canonical_url,
        )

        out_row = [
            r["sku"], r["category"], r["brand"], r["core"], r["spec"],
            r["stock"], r["price"], r["cleaned_title"],
            slug, canonical_url, canonical_slug, is_canonical,
            seo["en"]["title"], seo["en"]["description"], seo["en"]["keywords"],
            seo["bn"]["title"], seo["bn"]["description"], seo["bn"]["keywords"],
            jsonld,
            r["was_fixed"],
        ]
        ws_out.append(out_row)
        final_rows.append(out_row)
        dedupe_csv_rows.append((slug, canonical_slug))

        if (i + 1) % 1000 == 0:
            print(f"  step 3: {i + 1} rows built...")

    out_path = Path(args.out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(out_path)
    print(f"step 3: wrote {out_path} ({len(final_rows)} rows)")

    # ---------- dedupe map CSV ----------
    dedupe_map_path = Path(args.dedupe_map_path)
    with dedupe_map_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["slug", "canonical_slug", "is_canonical"])
        for slug, canon in dedupe_csv_rows:
            w.writerow([slug, canon, slug == canon])
    print(f"step 3: wrote {dedupe_map_path}")

    # ---------- sample CSV (first N rows, full schema) ----------
    sample_path = Path(args.sample_csv_path)
    n_sample = min(args.sample_rows, len(final_rows))
    with sample_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(list(_HEADER) + list(_NEW_COLS))
        for row in final_rows[:n_sample]:
            w.writerow(row)
    print(f"step 4: wrote {sample_path} ({n_sample} rows)")

    # ---------- summary ----------
    print()
    print("=" * 60)
    print(f"  rows processed       : {len(cleaned_rows)}")
    print(f"  audience tokens fixed: {n_audience_fixed} "
          f"({100.0 * n_audience_fixed / len(cleaned_rows):.1f}%)")
    print(f"  canonical URLs       : {n_canonical}")
    print(f"  redirect URLs        : {n_redirects} "
          f"({100.0 * n_redirects / len(cleaned_rows):.1f}%)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
