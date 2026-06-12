"""CSV/XLSX/TSV parser. Returns iterator of (row_number, dict) tuples.

Row number is 1-based and corresponds to the source-file line minus the
header row (data row 1 = file line 2). XLSX path uses openpyxl in
read-only mode so memory stays flat for 10k-row files.
"""
from __future__ import annotations

import csv
import io
from collections.abc import Iterator

from app.modules.bulk_upload.codes import (
    OPTIONAL_COLUMNS,
    REQUIRED_COLUMNS,
)

_ALL_KNOWN = set(REQUIRED_COLUMNS) | set(OPTIONAL_COLUMNS)

_ALIASES = {
    "price": "price_minor",
    "price_paisa": "price_minor",
    "stock": "stock_qty",
    "quantity": "stock_qty",
    "qty": "stock_qty",
    "name": "title",
    "category_slug": "category",
    "brand_name": "brand",
    "image": "image_url",
    "weight": "weight_grams",
}


def _normalise_header(raw: str) -> str:
    s = raw.strip().lower()
    # strip parenthesised hints like "Price (paisa)" → "price"
    if "(" in s:
        s = s.split("(", 1)[0].strip()
    s = s.replace(" ", "_").replace("-", "_")
    return _ALIASES.get(s, s)


def detect_columns(header_row: list[str]) -> dict[str, int]:
    """Map normalised column name → 0-based column index."""
    out: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        if raw is None:
            continue
        norm = _normalise_header(str(raw))
        if norm in _ALL_KNOWN and norm not in out:
            out[norm] = idx
    return out


def _row_to_dict(
    raw_row: list, col_map: dict[str, int],
) -> dict[str, str]:
    out: dict[str, str] = {}
    for col, idx in col_map.items():
        if idx < len(raw_row):
            val = raw_row[idx]
            out[col] = "" if val is None else str(val).strip()
        else:
            out[col] = ""
    return out


def parse_csv_or_tsv(
    file_bytes: bytes, *, delimiter: str = ",",
) -> Iterator[tuple[int, dict[str, str]]]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    try:
        header = next(reader)
    except StopIteration:
        return
    col_map = detect_columns(header)
    for line_idx, row in enumerate(reader, start=1):
        if not row or all((c is None or str(c).strip() == "") for c in row):
            continue
        yield line_idx, _row_to_dict(row, col_map)


def parse_xlsx(file_bytes: bytes) -> Iterator[tuple[int, dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise XlsxNotSupported(
            "xlsx not yet supported in this build — re-save as CSV.",
        ) from e
    wb = load_workbook(
        io.BytesIO(file_bytes), read_only=True, data_only=True,
    )
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows_iter))
        except StopIteration:
            return
        col_map = detect_columns([
            "" if c is None else str(c) for c in header
        ])
        for line_idx, row in enumerate(rows_iter, start=1):
            row_list = list(row)
            if not row_list or all(
                (c is None or str(c).strip() == "") for c in row_list
            ):
                continue
            yield line_idx, _row_to_dict(row_list, col_map)
    finally:
        wb.close()


class XlsxNotSupported(RuntimeError):
    """Raised when an XLSX file is uploaded but openpyxl is missing."""


def parse_file(
    file_bytes: bytes, file_format: str,
) -> Iterator[tuple[int, dict[str, str]]]:
    fmt = file_format.lower()
    if fmt == "csv":
        yield from parse_csv_or_tsv(file_bytes, delimiter=",")
    elif fmt == "tsv":
        yield from parse_csv_or_tsv(file_bytes, delimiter="\t")
    elif fmt == "xlsx":
        yield from parse_xlsx(file_bytes)
    else:
        raise ValueError(f"unsupported file_format: {file_format!r}")
