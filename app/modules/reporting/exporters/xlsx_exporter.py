"""XLSX exporter using openpyxl.

Single-sheet workbook with:
  - title in row 1 (merged across all columns)
  - generated-at + row-count in row 2
  - column headers in row 3 (frozen pane)
  - data rows from row 4 onward

Bengali text renders fine in Excel because XLSX cells are stored as
UTF-8 inside the zip — no font registration required (Excel falls
back to system fonts at display time).
"""

from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.modules.reporting.exporters._format import format_cell

_HEADER_FILL = PatternFill(start_color="FFEEEEEE", end_color="FFEEEEEE", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=14)
_META_FONT = Font(italic=True, size=9)


def render(
    *,
    title: str,
    columns: list[dict[str, str]],
    rows: list[dict[str, Any]],
    generated_at: datetime,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    # 31-char limit on Excel sheet names — truncate defensively.
    ws.title = (title or "Report")[:31]

    n_cols = max(1, len(columns))

    # --- title row ---
    ws.cell(row=1, column=1, value=title).font = _TITLE_FONT
    ws.merge_cells(
        start_row=1, start_column=1,
        end_row=1, end_column=n_cols,
    )

    # --- meta row ---
    meta = (
        f"Generated {generated_at.isoformat(timespec='seconds')} "
        f"• {len(rows)} row(s)"
    )
    ws.cell(row=2, column=1, value=meta).font = _META_FONT
    ws.merge_cells(
        start_row=2, start_column=1,
        end_row=2, end_column=n_cols,
    )

    # --- header row ---
    for ci, col in enumerate(columns, start=1):
        cell = ws.cell(
            row=3, column=ci,
            value=col.get("label") or col.get("key", ""),
        )
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    ws.freeze_panes = "A4"

    # --- data rows ---
    for ri, r in enumerate(rows, start=4):
        for ci, col in enumerate(columns, start=1):
            raw = r.get(col.get("key", ""))
            ctype = col.get("type")
            # Numeric types stay native so Excel can sum/sort them.
            if ctype in ("int", "money", "decimal"):
                ws.cell(row=ri, column=ci, value=_to_number(raw))
            else:
                ws.cell(row=ri, column=ci, value=format_cell(raw, ctype))

    # --- column widths (rough auto-fit) ---
    for ci, col in enumerate(columns, start=1):
        header = col.get("label") or col.get("key", "")
        max_data_len = max(
            (len(format_cell(r.get(col.get("key", "")), col.get("type")))
             for r in rows[:200]),  # sample first 200 to bound cost
            default=0,
        )
        # 8 = baseline, +2 padding; cap at 60 to avoid silly wide columns.
        width = min(60, max(8, max(len(header), max_data_len) + 2))
        ws.column_dimensions[get_column_letter(ci)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _to_number(v: Any) -> float | int | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int | float):
        return v
    if isinstance(v, Decimal):
        # Decimal → float at display only; the underlying string in
        # the cell is fine for downstream readers.
        return float(v)
    try:
        return float(v)
    except (TypeError, ValueError):
        return None
